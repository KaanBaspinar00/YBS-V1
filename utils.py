# utils.py

import os
import qrcode
import sqlite3
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
import uuid
from io import BytesIO
from pushbullet import Pushbullet
import streamlit as st
import json
# ---------------------- File Paths ----------------------

# File paths
main_excel_file = "depo_veri12.xlsx"
qr_images_file = "qr_images12.xlsx"
recent_qr_codes_file = "recent_qr_codes12.xlsx"
usage_log_file = "usage_log12.xlsx"
db_file = "depo_veri12.db"
qr_codes_folder = "qr_codes12"
asset_movements_excel = "asset_movements.xlsx"
qr_codes_output_excel = "qr_codes_output.xlsx"

# Ensure the QR codes folder exists
os.makedirs(qr_codes_folder, exist_ok=True)


# ---------------------- Database Initialization ----------------------

def init_db():
    """Initialize SQLite database and create tables if they don't exist."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()

    # Create varliklar table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS varliklar (
            entry_id INTEGER PRIMARY KEY AUTOINCREMENT,
            id TEXT UNIQUE,
            qr_kodu TEXT,
            varlık_adı TEXT,
            gönderen TEXT,
            alıcı TEXT,
            miktar INTEGER,
            unit TEXT,
            adet INTEGER,
            kacıncı INTEGER,
            zaman TEXT,
            quantity INTEGER DEFAULT 0
        )
    ''')

    # Create asset_movements table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS asset_movements (
            movement_id INTEGER PRIMARY KEY AUTOINCREMENT,
            asset_id TEXT,
            action TEXT,
            quantity INTEGER,
            partner_firm TEXT,
            timestamp TEXT,
            worker TEXT,
            notes TEXT,
            FOREIGN KEY (asset_id) REFERENCES varliklar (id)
        )
    ''')

    # Create tasks table
    cursor.execute('''
            CREATE TABLE IF NOT EXISTS tasks (
                task_id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT,
                description TEXT,
                assigned_to TEXT,  -- Store as comma-separated usernames
                created_by TEXT,
                urgency INTEGER,
                status TEXT DEFAULT 'not seen',
                progress INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP
            )
        ''')

    conn.commit()
    conn.close()


# Call init_db to initialize the database when utils.py is imported
init_db()


# ---------------------- Excel File Initialization ----------------------

def create_excel_file_if_missing(file_path, headers):
    """Create an Excel file with specified headers if it doesn't exist."""
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(file_path)
        print(f"Created missing Excel file: {file_path}")


# Check and create necessary Excel files with headers
create_excel_file_if_missing(main_excel_file,
                             ["id", "gönderen", "alıcı", "varlık_adı", "miktar", "unit", "quantity", "zaman"])
create_excel_file_if_missing(qr_images_file, ["id", "QR-codes-text", "QR-codes-images"])
create_excel_file_if_missing(recent_qr_codes_file, ["id", "QR-codes-text", "QR-codes-images"])
create_excel_file_if_missing(usage_log_file, ["log_id", "asset_id", "origin", "destination", "timestamp"])
create_excel_file_if_missing(asset_movements_excel,
                             ["Zaman", "Varlık ID", "Varlık Adı", "Aksiyon", "Miktar", "Firma", "Çalışan", "Notlar"])


# ---------------------- Utility Functions ----------------------

def create_empty_excel(file_path, headers):
    """Create an empty Excel file with specified headers."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    wb.save(file_path)


def parse_qr_code_data(qr_code_text):
    """Parse the QR code text to extract asset information."""
    try:
        # Split the qr_code_text by dashes
        parts = qr_code_text.split('-', 7)

        # Ensure the main parts are present
        assetname = parts[0]
        gönderen = parts[1]
        alıcı = parts[2]
        miktar_unit = parts[3]
        adet = int(parts[4])
        kacıncı = int(parts[5])
        zaman = parts[6]

        # Remaining part is UUID (if present)
        uuid_part = parts[7] if len(parts) > 7 else None

        # Process miktar and unit
        miktar = ''.join(filter(str.isdigit, miktar_unit))
        unit = ''.join(filter(str.isalpha, miktar_unit))

        return {
            "assetname": assetname,
            "gönderen": gönderen,
            "alıcı": alıcı,
            "miktar": int(miktar),
            "unit": unit,
            "adet": adet,
            "kacıncı": kacıncı,
            "zaman": zaman,
            "uuid": uuid_part
        }
    except Exception as e:
        print(f"Error parsing QR code data: {e}")
        return None


# ---------------------- Database Functions ----------------------

def add_asset_to_db(id_, qr_kodu, varlık_adı, gönderen, alıcı, miktar, unit, adet, kacıncı, zaman):
    """Add a new asset entry to the SQLite database."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO varliklar (id, qr_kodu, varlık_adı, gönderen, alıcı, miktar, unit, adet, kacıncı, zaman, quantity)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (id_, qr_kodu, varlık_adı, gönderen, alıcı, miktar, unit, adet, kacıncı, zaman, adet))
        conn.commit()
        print(f"Asset {id_} successfully added to the database.")
    except sqlite3.IntegrityError as e:
        print(f"Error adding asset {id_} to database: {e}")
    finally:
        conn.close()

    title = "Yeni Varlık Eklendi"
    message = f"Varlık Adı: {varlık_adı}\nID: {id_}\nMiktar: {miktar}{unit}\nAdet: {adet}"
    send_pushbullet_notification(title, message)


def get_asset_by_qr(qr_code_text):
    """Fetch asset information from the database using the QR code text."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM varliklar WHERE id = ?", (qr_code_text,))
    asset = cursor.fetchone()
    conn.close()
    return asset


def get_asset_by_id(asset_id):
    """Fetch asset information from the database using the asset ID."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM varliklar WHERE id = ?", (asset_id,))
    asset = cursor.fetchone()
    conn.close()
    return asset


def get_asset_by_name(varlik_adi):
    """Fetch asset information from the database using the asset name."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM varliklar WHERE varlık_adı = ?", (varlik_adi,))
    asset = cursor.fetchone()
    conn.close()
    return asset


def decrease_asset_quantity(asset_id):
    """Decrease the quantity of an asset by 1."""
    update_asset_quantity(asset_id, -1)


def undo_decrease_asset_quantity(asset_id):
    """Increase the quantity of an asset by 1."""
    update_asset_quantity(asset_id, 1)


def update_asset_quantity(asset_id, delta_quantity):
    """Update the asset's quantity by delta."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("UPDATE varliklar SET quantity = quantity + ? WHERE id = ?", (delta_quantity, asset_id))
    conn.commit()
    conn.close()


def delete_asset_from_db(varlik_id):
    """Delete an asset from the database."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM varliklar WHERE id = ?", (varlik_id,))
    conn.commit()
    conn.close()


def add_usage_log(asset_id, origin, destination, timestamp):
    """Add a usage log entry to the SQLite database."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO usage_logs (asset_id, origin, destination, timestamp)
        VALUES (?, ?, ?, ?)
    ''', (asset_id, origin, destination, timestamp))
    conn.commit()
    conn.close()


def log_asset_movement(asset_id, action, quantity, partner_firm, worker, notes):
    """Log an asset movement."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute('''
        INSERT INTO asset_movements (asset_id, action, quantity, partner_firm, timestamp, worker, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (asset_id, action, quantity, partner_firm, timestamp, worker, notes))
    movement_id = cursor.lastrowid  # Get the ID of the inserted movement
    conn.commit()
    conn.close()

    # Update the asset quantity in varliklar table
    if action == 'used' or action == 'sent_for_processing':
        update_asset_quantity(asset_id, -quantity)
    elif action == 'received_back':
        update_asset_quantity(asset_id, quantity)

    # Update Excel files
    update_main_excel(asset_id)
    update_asset_movements_excel()

    # Send notification to admins
    action_map = {
        'used': 'Kullanıldı',
        'sent_for_processing': 'İşlem İçin Gönderildi',
        'received_back': 'Geri Alındı'
    }
    action_text = action_map.get(action, action)
    title = f"Çalışan Aksiyonu: {action_text}"
    message = f"Varlık ID: {asset_id}\nMiktar: {quantity}\nÇalışan: {worker}"
    if partner_firm:
        message += f"\nFirma: {partner_firm}"
    if notes:
        message += f"\nNotlar: {notes}"

    send_pushbullet_notification(title, message)

    return movement_id  # Return the movement ID


def undo_asset_movement(movement_id):
    """Undo a logged asset movement."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    # Fetch the movement details
    cursor.execute("SELECT asset_id, action, quantity FROM asset_movements WHERE movement_id = ?", (movement_id,))
    movement = cursor.fetchone()
    if movement:
        asset_id, action, quantity = movement
        # Reverse the quantity change
        if action == 'used' or action == 'sent_for_processing':
            update_asset_quantity(asset_id, quantity)
        elif action == 'received_back':
            update_asset_quantity(asset_id, -quantity)
        # Delete the movement record
        cursor.execute("DELETE FROM asset_movements WHERE movement_id = ?", (movement_id,))
        conn.commit()
        conn.close()
        # Update Excel files
        update_main_excel(asset_id)
        update_asset_movements_excel()
        return True
    conn.close()
    return False

def get_asset_history():
    """Fetch the latest 10 asset entries from the database."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, varlık_adı, gönderen, alıcı, zaman
        FROM varliklar
        ORDER BY zaman DESC
        LIMIT 10
    """)
    rows = cursor.fetchall()
    conn.close()
    return rows


def get_current_stock_levels():
    """Return a DataFrame with current stock levels."""
    conn = sqlite3.connect(db_file)
    df = pd.read_sql_query("SELECT id, varlık_adı, gönderen, alıcı, miktar, unit, quantity FROM varliklar", conn)
    conn.close()
    return df


def get_asset_movements():
    """Return a DataFrame with asset movement logs."""
    conn = sqlite3.connect(db_file)
    df = pd.read_sql_query("""
        SELECT
            am.timestamp AS 'Zaman',
            am.asset_id AS 'Varlık ID',
            v.varlık_adı AS 'Varlık Adı',
            CASE
                WHEN am.action = 'used' THEN 'Kullanıldı'
                WHEN am.action = 'sent_for_processing' THEN 'İşlem İçin Gönderildi'
                WHEN am.action = 'received_back' THEN 'Geri Alındı'
                ELSE am.action
            END AS 'Aksiyon',
            am.quantity AS 'Miktar',
            am.partner_firm AS 'Firma',
            am.worker AS 'Çalışan',
            am.notes AS 'Notlar'
        FROM asset_movements am
        JOIN varliklar v ON am.asset_id = v.id
        ORDER BY am.timestamp DESC
    """, conn)
    conn.close()
    return df


def get_asset_history_by_id(asset_id):
    """Return a DataFrame with the movement history of a specific asset."""
    conn = sqlite3.connect(db_file)
    df = pd.read_sql_query("""
        SELECT
            am.timestamp AS 'Zaman',
            CASE
                WHEN am.action = 'used' THEN 'Kullanıldı'
                WHEN am.action = 'sent_for_processing' THEN 'İşlem İçin Gönderildi'
                WHEN am.action = 'received_back' THEN 'Geri Alındı'
                ELSE am.action
            END AS 'Aksiyon',
            am.quantity AS 'Miktar',
            am.partner_firm AS 'Firma',
            am.worker AS 'Çalışan',
            am.notes AS 'Notlar'
        FROM asset_movements am
        WHERE am.asset_id = ?
        ORDER BY am.timestamp DESC
    """, conn, params=(asset_id,))
    conn.close()
    return df


def get_filtered_assets(filter_option, filter_query):
    """Return a DataFrame of assets based on filter criteria."""
    conn = sqlite3.connect(db_file)
    query = "SELECT id, varlık_adı, gönderen, alıcı, miktar, unit, quantity FROM varliklar"
    params = ()
    if filter_option != "Hepsi" and filter_query:
        column_map = {
            "Varlık Adı": "varlık_adı",
            "Gönderen": "gönderen",
            "Alıcı": "alıcı"
        }
        column_name = column_map.get(filter_option)
        query += f" WHERE {column_name} LIKE ?"
        params = ('%' + filter_query + '%',)
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df


def get_filtered_movements(action_type, start_date, end_date):
    """Return a DataFrame of asset movements based on filter criteria."""
    conn = sqlite3.connect(db_file)
    query = """
        SELECT
            am.timestamp AS 'Zaman',
            am.asset_id AS 'Varlık ID',
            v.varlık_adı AS 'Varlık Adı',
            CASE
                WHEN am.action = 'used' THEN 'Kullanıldı'
                WHEN am.action = 'sent_for_processing' THEN 'İşlem İçin Gönderildi'
                WHEN am.action = 'received_back' THEN 'Geri Alındı'
                ELSE am.action
            END AS 'Aksiyon',
            am.quantity AS 'Miktar',
            am.partner_firm AS 'Firma',
            am.worker AS 'Çalışan',
            am.notes AS 'Notlar'
        FROM asset_movements am
        JOIN varliklar v ON am.asset_id = v.id
        WHERE 1=1
    """
    params = []
    if action_type != "Hepsi":
        action_map = {
            "Kullanıldı": "used",
            "İşlem İçin Gönderildi": "sent_for_processing",
            "Geri Alındı": "received_back"
        }
        action_code = action_map.get(action_type)
        query += " AND am.action = ?"
        params.append(action_code)
    if start_date and end_date:
        query += " AND DATE(am.timestamp) BETWEEN ? AND ?"
        params.extend([start_date, end_date])
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df


# ---------------------- Excel Functions ----------------------

def append_to_qr_excel(qr_code_text, qr_image_path, excel_path, id_=None):
    """Append a new row to the specified Excel file with id, QR code text, and image."""
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "QR-codes-text", "QR-codes-images"])
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1, value=id_)
    ws.cell(row=new_row, column=2, value=qr_code_text)

    # Insert QR code image
    img = XLImage(qr_image_path)
    img.width, img.height = 100, 100
    ws.add_image(img, f"C{new_row}")
    wb.save(excel_path)


def append_to_main_excel(id_, varlık_adı, gönderen, alıcı, miktar, unit, adet, kacıncı, zaman):
    """Append or update asset information in the main Excel file."""
    if os.path.exists(main_excel_file):
        wb = load_workbook(main_excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        headers = ["id", "gönderen", "alıcı", "varlık_adı", "miktar", "unit", "quantity", "zaman"]
        ws.append(headers)

    # Check if asset exists
    ids = [cell.value for cell in ws['A'] if cell.value != 'id']
    if id_ in ids:
        # Update existing asset
        for row in ws.iter_rows(min_row=2):
            if row[0].value == id_:
                row[6].value = adet  # Update quantity
    else:
        # Append new asset
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=id_)
        ws.cell(row=new_row, column=2, value=gönderen)
        ws.cell(row=new_row, column=3, value=alıcı)
        ws.cell(row=new_row, column=4, value=varlık_adı)
        ws.cell(row=new_row, column=5, value=miktar)
        ws.cell(row=new_row, column=6, value=unit)
        ws.cell(row=new_row, column=7, value=adet)  # quantity
        ws.cell(row=new_row, column=8, value=zaman)

    # Apply formatting
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = ws['A2']
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    ws.conditional_formatting.add(f'G2:G{ws.max_row}', CellIsRule(operator='lessThan', formula=['5'], fill=red_fill))
    wb.save(main_excel_file)


def update_main_excel(asset_id):
    """Update the main Excel file to reflect current stock levels."""
    # Fetch asset data from database
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM varliklar WHERE id = ?", (asset_id,))
    asset = cursor.fetchone()
    conn.close()

    if asset:
        entry_id, id_, qr_kodu, varlık_adı, gönderen, alıcı, miktar, unit, adet, kacıncı, zaman, quantity = asset
        append_to_main_excel(id_, varlık_adı, gönderen, alıcı, miktar, unit, quantity, kacıncı, zaman)


def update_asset_movements_excel():
    """Update the asset movements Excel file."""
    movements_data = get_asset_movements()
    movements_data.to_excel(asset_movements_excel, index=False)


def append_to_usage_log_excel(asset_id, origin, destination, timestamp, excel_path=usage_log_file):
    """Append a new row to the usage log Excel file."""
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["log_id", "asset_id", "origin", "destination", "timestamp"])
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1, value=new_row - 1)  # log_id
    ws.cell(row=new_row, column=2, value=asset_id)
    ws.cell(row=new_row, column=3, value=origin)
    ws.cell(row=new_row, column=4, value=destination)
    ws.cell(row=new_row, column=5, value=timestamp)
    wb.save(excel_path)


def convert_df_to_excel(df):
    """Convert a DataFrame to an Excel file in memory."""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False)
    writer.close()
    processed_data = output.getvalue()
    return processed_data


# ---------------------- QR Code Generation ----------------------

def generate_qr_codes(assetname, unit, miktar, adet, gönderen, alıcı):
    """Generate QR codes based on the input parameters and update Excel and DB."""
    zaman = datetime.now().strftime("%Y-%m-%d-%H-%M-%S-%f")
    qr_codes_info = []

    for kacıncı in range(1, adet + 1):
        uuid_part = uuid.uuid4().hex
        unique_id = f"{assetname}-{gönderen}-{alıcı}-{miktar}{unit}-{adet}-{kacıncı}-{zaman}-{uuid_part}"
        qr_code_text = unique_id

        # Generate and save the QR code image
        qr = qrcode.make(qr_code_text)
        qr_image_path = os.path.join(qr_codes_folder, f"{unique_id}.png")
        qr.save(qr_image_path)

        # Append to QR Excel files
        append_to_qr_excel(qr_code_text, qr_image_path, qr_images_file, id_=unique_id)
        append_to_qr_excel(qr_code_text, qr_image_path, recent_qr_codes_file, id_=unique_id)

        # Add to SQLite Database
        add_asset_to_db(unique_id, qr_code_text, assetname, gönderen, alıcı, miktar, unit, adet, kacıncı, zaman)

        # Append to main Excel
        append_to_main_excel(unique_id, assetname, gönderen, alıcı, miktar, unit, adet, kacıncı, zaman)

        # Collect info for Excel with QR codes
        qr_codes_info.append({'id': unique_id, 'image_path': qr_image_path})

    # After generating all QR codes, create the Excel file with QR codes
    create_excel_with_qr_codes(qr_codes_info, qr_codes_output_excel)

    return qr_codes_info


def create_excel_with_qr_codes(qr_codes_info, excel_file_path):
    """Create an Excel file containing QR code images and their IDs with proper formatting."""
    wb = Workbook()
    ws = wb.active

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40

    # Set headers
    ws.cell(row=1, column=1, value='QR Kodları')
    ws.cell(row=1, column=2, value='ID')
    ws.row_dimensions[1].height = 30

    # Style headers
    header_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = header_font

    row_num = 2
    for qr_info in qr_codes_info:
        id_ = qr_info['id']
        image_path = qr_info['image_path']

        # Insert QR code image
        img = XLImage(image_path)
        img.width, img.height = 150, 150
        img_cell = f"A{row_num}"
        ws.add_image(img, img_cell)
        ws.row_dimensions[row_num].height = 110  # Adjust row height for image

        # Write ID next to the image
        ws.cell(row=row_num, column=2, value=id_)
        ws.row_dimensions[row_num].height = 110  # Ensure row height matches

        # Move to the next row
        row_num += 1

    wb.save(excel_file_path)


def get_recent_qr_code_images():
    """Get file paths of recently generated QR code images."""
    if os.path.exists(recent_qr_codes_file):
        df_recent = pd.read_excel(recent_qr_codes_file)
        image_paths = []
        for index, row in df_recent.iterrows():
            image_path = os.path.join(qr_codes_folder, f"{row['id']}.png")
            if os.path.exists(image_path):
                image_paths.append(image_path)
        return image_paths
    return []


# ---------------------- Additional Utility Functions ----------------------

def add_assets_from_recent_qr_codes():
    """Add assets from recent QR codes Excel file to main Excel and database."""
    if os.path.exists(recent_qr_codes_file):
        df_recent = pd.read_excel(recent_qr_codes_file)
        for index, row in df_recent.iterrows():
            id_ = row["id"]
            qr_code_text = row["QR-codes-text"]
            parts = qr_code_text.split('-', 7)
            assetname, gönderen, alıcı, miktar_unit, adet, kacıncı, zaman, uuid_part = parts
            miktar = ''.join(filter(str.isdigit, miktar_unit))
            unit = ''.join(filter(str.isalpha, miktar_unit))
            append_to_main_excel(id_, assetname, gönderen, alıcı, int(miktar), unit, int(adet), int(kacıncı), zaman)


def delete_asset(varlik_id):
    """Delete an asset from the database and relevant Excel files."""
    # Delete movements related to the asset
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM asset_movements WHERE asset_id = ?", (varlik_id,))
    conn.commit()
    conn.close()

    # Delete from main_excel_file
    if os.path.exists(main_excel_file):
        df = pd.read_excel(main_excel_file)
        df = df[df["id"] != varlik_id]
        df.to_excel(main_excel_file, index=False)

    # Delete from qr_images_file
    if os.path.exists(qr_images_file):
        wb_qr = load_workbook(qr_images_file)
        ws_qr = wb_qr.active
        rows_to_delete = []
        for row in ws_qr.iter_rows(min_row=2, values_only=False):
            if row[0].value == varlik_id:
                rows_to_delete.append(row[0].row)
        for row_num in reversed(rows_to_delete):  # Delete from bottom to top
            ws_qr.delete_rows(row_num, 1)
        wb_qr.save(qr_images_file)

    # Delete from recent_qr_codes_file
    if os.path.exists(recent_qr_codes_file):
        wb_recent_qr = load_workbook(recent_qr_codes_file)
        ws_recent_qr = wb_recent_qr.active
        rows_to_delete_recent = []
        for row in ws_recent_qr.iter_rows(min_row=2, values_only=False):
            if row[0].value == varlik_id:
                rows_to_delete_recent.append(row[0].row)
        for row_num in reversed(rows_to_delete_recent):
            ws_recent_qr.delete_rows(row_num, 1)
        wb_recent_qr.save(recent_qr_codes_file)

    # Delete QR code image
    qr_image_path = os.path.join(qr_codes_folder, f"{varlik_id}.png")
    if os.path.exists(qr_image_path):
        os.remove(qr_image_path)

    # Send notification to admins
    title = "Admin Aksiyonu: Varlık Silindi"
    message = f"Varlık ID: {varlik_id} silindi.\nSilme işlemini yapan: {st.session_state.username}"
    send_pushbullet_notification(title, message)


def get_assets_from_main_excel():
    """Fetch all assets from the main Excel file."""
    if os.path.exists(main_excel_file):
        return pd.read_excel(main_excel_file)
    return pd.DataFrame()

def get_last_movement_by_worker(worker_name):
    """Get the last movement made by the worker."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT movement_id, asset_id, action, quantity
        FROM asset_movements
        WHERE worker = ?
        ORDER BY timestamp DESC
        LIMIT 1
    """, (worker_name,))
    last_movement = cursor.fetchone()
    conn.close()
    return last_movement

# ---------------------- User Authentication ----------------------

users = {
    'yılmaz': {'password': '0', 'role': 'admin'},
    'canan': {'password': '0', 'role': 'admin'},
    'sinan': {'password': '0', 'role': 'admin'},
    'Mert': {'password': '0', 'role': 'worker'},
    'Umut': {'password': '0', 'role': 'worker'},
    'Ceren': {'password': '0', 'role': 'worker'}
    # Add more users as needed
}

def authenticate_user(username, password):
    """Authenticate user and return their role."""
    user = users.get(username)
    if user and user['password'] == password:
        return user['role']
    return None

# Add this function to send notifications
def send_pushbullet_notification(title, message):
    """Send a notification via Pushbullet."""
    # Get the API key from secrets
    if "PUSHBULLET_API_KEY" in st.secrets:
        api_key = st.secrets["PUSHBULLET_API_KEY"]
    else:
        print("Pushbullet API key not found in secrets.")
        return

    try:
        pb = Pushbullet(api_key)
        pb.push_note(title, message)
        print(f"Notification sent: {title} - {message}")
    except Exception as e:
        print(f"Failed to send Pushbullet notification: {e}")


def create_task(title, description, assigned_to, urgency):
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    assigned_to_str = ",".join(assigned_to)  # Convert list of users to comma-separated string
    cursor.execute('''
        INSERT INTO tasks (title, description, assigned_to, created_by, urgency, status)
        VALUES (?, ?, ?, ?, ?, 'not seen')
    ''', (title, description, assigned_to_str, st.session_state.username, urgency))
    conn.commit()
    task_id = cursor.lastrowid  # Get the ID of the newly created task
    conn.close()
    return task_id


def update_task_progress(task_id, progress):
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute('UPDATE tasks SET progress = ?, updated_at = CURRENT_TIMESTAMP WHERE task_id = ?', (progress, task_id))
    conn.commit()
    conn.close()

def complete_task(task_id):
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute('UPDATE tasks SET status = "completed", progress = 100, updated_at = CURRENT_TIMESTAMP WHERE task_id = ?', (task_id,))
    conn.commit()
    conn.close()

#################################################################################################
def get_task_history():
    """Fetch recent task history for display in the sidebar."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT task_id, title, status, progress, created_at
        FROM tasks
        ORDER BY created_at DESC
        LIMIT 10
    """)
    tasks = cursor.fetchall()
    conn.close()
    return [dict(zip([column[0] for column in cursor.description], task)) for task in tasks]

def get_admin_usernames():
    """Fetch admin usernames from roles.txt."""
    try:
        with open("roles.txt", "r") as file:
            user_data = json.load(file)  # Load data as JSON

        # Filter for users with the role 'admin'
        admins = [username for username, details in user_data['users'].items() if details['role'] == 'admin']
        return admins
    except json.JSONDecodeError:
        print("Error reading admin usernames from roles.txt: The file format is incorrect. Please ensure it is valid JSON.")
        return []
    except FileNotFoundError:
        print("Error: roles.txt file not found. Please ensure the file is in the correct directory.")
        return []


def get_open_tasks():
    """Fetch open tasks for display, ordered by urgency and creation date."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT * FROM tasks
        WHERE status != 'completed'
        ORDER BY urgency DESC, created_at ASC
    """)
    tasks = cursor.fetchall()
    conn.close()
    return [dict(zip([column[0] for column in cursor.description], task)) for task in tasks]

def get_task_creator(task_id):
    """Fetch the creator's username for a given task ID."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("SELECT created_by FROM tasks WHERE task_id = ?", (task_id,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else None  # Return the username or None if not found


def mark_task_as_seen(task_id, viewer):
    """Mark a task as seen, update its status, and notify the task creator."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute('UPDATE tasks SET status = "in progress", updated_at = CURRENT_TIMESTAMP WHERE task_id = ?', (task_id,))
    conn.commit()

    # Add notification to the creator
    creator = get_task_creator(task_id)

    conn.close()

def get_completed_tasks():
    """Fetch completed tasks for display, ordered by completion date."""
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT * FROM tasks
        WHERE status = 'completed'
        ORDER BY updated_at DESC
    """)
    tasks = cursor.fetchall()
    conn.close()
    return [dict(zip([column[0] for column in cursor.description], task)) for task in tasks]
